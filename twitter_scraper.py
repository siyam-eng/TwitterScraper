import tweepy
import webbrowser
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import json
import requests
from requests import Session
import random
import time


# Dealing with Excel Files
FILE_PATH = "Twitter Data Error.xlsx"
wb = load_workbook(FILE_PATH)

# Consumer API keys
with open("config.json") as f:
    CONFIG_DATA = json.load(f)
consumer_key = CONFIG_DATA.get("consumer_key")
consumer_secret_key = CONFIG_DATA.get("consumer_secret_key")
callback_uri = CONFIG_DATA.get("callback_uri")

# Authenticating the app
auth = tweepy.OAuthHandler(consumer_key, consumer_secret_key, callback_uri)
redirect_url = auth.get_authorization_url()

# opening the link with browser
webbrowser.open(redirect_url)
user_pin_input = input("What is your pin? ")
auth.get_access_token(user_pin_input)

# calling the api
api = tweepy.API(auth)

# list of user-agents to be used in request headers
HEADERS_LIST = [
    "Mozilla/5.0 (Windows; U; Windows NT 6.1; x64; fr; rv:1.9.2.13) Gecko/20101203 Firebird/3.6.13",
    "Mozilla/5.0 (compatible, MSIE 11, Windows NT 6.3; Trident/7.0; rv:11.0) like Gecko",
    "Mozilla/5.0 (Windows; U; Windows NT 6.1; rv:2.2) Gecko/20110201",
    "Opera/9.80 (X11; Linux i686; Ubuntu/14.10) Presto/2.12.388 Version/12.16",
    "Mozilla/5.0 (Windows NT 5.2; RW; rv:7.0a1) Gecko/20091211 SeaMonkey/9.23a1pre",
]


# takes an url and returns its status code and final redirected url
def get_response_code(url):
    global wb
    if url:
        # creating the request Session
        session = Session()
        header = {
            "User-Agent": random.choice(HEADERS_LIST),
            "X-Requested-With": "XMLHttpRequest",
        }
        session.headers.update(header)

        url = "https://" + url if not url.startswith("http") else url
        final_url = url
        try:
            response = session.get(url)
            final_url = response.url
        except requests.exceptions.SSLError:
            response = session.get(url, verify=False)
            final_url = response.url
        except requests.exceptions.ConnectionError as err:
            final_url = url
            error = str(err)
            wb["Errors"].append(
                ("Failed to get the final url for ", url, "Due to", error)
            )
        except Exception as err:
            wb["Errors"].append(
                ("Failed to get the final url for ", url, "Due to", str(err))
            )
        return final_url


# fetching the user
def get_user_data(screen_name):
    user = api.get_user(screen_name)

    # fetching the description
    name = user.name
    followers_count = user.followers_count
    following_count = user.friends_count
    location = user.location
    verified = user.verified
    description = user.description
    created_at = user.created_at.strftime("%d %B %Y")
    statuses_count = user.statuses_count
    status_available = bool(statuses_count)
    url = get_response_code(user.url)

    data_dict = {
        "name": name,
        "followers_count": followers_count,
        "following_count": following_count,
        "description": description,
        "location": location,
        "url": url,
        "created_at": created_at,
        "verified": verified,
        "status_available": status_available,
        "statuses_count": statuses_count,
    }

    return data_dict


# Prepares the excel sheets and names the columns
def customize_excel_sheet():
    global wb
    output = (
        wb.create_sheet("Output") if "Output" not in wb.sheetnames else wb["Output"]
    )
    errors = (
        wb.create_sheet("Errors") if "Errors" not in wb.sheetnames else wb["Errors"]
    )

    font = Font(color="000000", bold=True)
    bg_color = PatternFill(fgColor="E8E8E8", fill_type="solid")

    # editing the output sheet
    output_column = zip(
        ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"),
        (
            "Screen Name",
            "Name / ID",
            "Followers",
            "Following",
            "Description",
            "Location",
            "Link",
            "Joined Date",
            "Verified",
            "Status Available",
            "Statuses Count",
        ),
    )
    for col, value in output_column:
        cell = output[f"{col}1"]
        cell.value = value
        cell.font = font
        cell.fill = bg_color
        output.freeze_panes = cell

        # fixing the column width
        output.column_dimensions[col].width = 20

    # save the customized file
    wb.save(FILE_PATH)


# Generates the input links
def generate_screen_names():
    global wb
    inputs = wb["Input"]
    for row in range(2, inputs.max_row + 1):
        # generates the links one by one
        if value := inputs[f"A{row}"].value:
            yield value


def insert_data_into_excel():
    customize_excel_sheet()
    for screen_name in generate_screen_names():
        print(f'Scraping {screen_name}')
        try:
            user_data = get_user_data(screen_name)

            wb["Output"].append(
                (
                    screen_name,
                    user_data["name"],
                    user_data["followers_count"],
                    user_data["following_count"],
                    user_data["description"],
                    user_data["location"],
                    user_data["url"],
                    user_data["created_at"],
                    user_data["verified"],
                    user_data["status_available"],
                    user_data["statuses_count"]
                )
            )
        except tweepy.error.TweepError as err:
            wb["Errors"].append((screen_name, str(err)))
        except Exception as excep:
            print(screen_name, excep)
            wb["Errors"].append((screen_name, str(excep)))

        # Take some rest 
        time.sleep(0.25)

        # save the file after each user 
        wb.save(FILE_PATH)

        # print completion statement
        print(f"THE SCRIPT RAN SUCCESSFULLY!")

    


# Calling the main function
insert_data_into_excel()




